"""
Functional Hazard Assessment (FHA) Excel Generator
Generates FHA worksheets from SysML model and populates template.
"""

from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, field, asdict
import json

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, 
    numbers
)
from openpyxl.utils import get_column_letter

import syside


# ============================================================================
# FHA Data Classes
# ============================================================================

@dataclass
class SystemMode:
    """Represents an operational system mode."""
    mode_id: str
    mode_name: str
    description: str
    entry_conditions: str = ""
    exit_conditions: str = ""


@dataclass
class FunctionalFailure:
    """Represents a functional failure for FHA."""
    fha_id: str
    mode: str
    system: str
    function_name: str
    failure_mode: str
    failure_description: str
    mode_phase: str
    hazard_id: Optional[str] = None
    hazard_effect: str = ""
    mishap: str = ""
    mitigations: str = ""
    severity: str = ""
    likelihood: str = ""
    risk: str = ""
    swci: str = ""  # Safety & Criticality Item
    scc_condition: str = ""
    scc_rationale: str = ""
    severity_condition: str = ""
    reference_requirements: str = ""
    assumptions: str = ""
    notes: str = ""


# ============================================================================
# SysML Model Extensions
# ============================================================================

class SysMLModelExtender:
    """Extends the SysML model with FHA-related definitions."""
    
    SYSTEM_MODES_PACKAGE = """
package SystemModes {
    doc /* System operational modes for FHA */
    private import GARC_Model::*;
    
    // System mode metadata
    requirement def SystemMode {
        attribute modeId : String;
        attribute modeName : String;
        attribute description : String;
        attribute entryConditions : String;
        attribute exitConditions : String;
    }
    
    // Define operational modes
    requirement MODE_001 : SystemMode {
        doc /* Startup and initialization mode */
        attribute modeId :>> modeId = "MODE-001";
        attribute modeName :>> modeName = "Startup & Initialization";
        attribute description :>> description = "System boots, performs self-tests, initializes sensors";
        attribute entryConditions :>> entryConditions = "Power applied, startup command";
        attribute exitConditions :>> exitConditions = "All systems initialized, ready for nominal operation";
    }
    
    requirement MODE_002 : SystemMode {
        doc /* Navigation and transit mode */
        attribute modeId :>> modeId = "MODE-002";
        attribute modeName :>> modeName = "Navigation & Transit";
        attribute description :>> description = "Vehicle navigates to target waypoint without payload arming";
        attribute entryConditions :>> entryConditions = "Initialization complete, navigation enabled";
        attribute exitConditions :>> exitConditions = "Reached target area or abort commanded";
    }
    
    requirement MODE_003 : SystemMode {
        doc /* Payload preparation and arming mode */
        attribute modeId :>> modeId = "MODE-003";
        attribute modeName :>> modeName = "Payload Operations";
        attribute description :>> description = "Payload system armed and ready for actuation";
        attribute entryConditions :>> entryConditions = "Correct position, operator authorization";
        attribute exitConditions :>> exitConditions = "Payload actuated or disarmed, exit operation";
    }
    
    requirement MODE_004 : SystemMode {
        doc /* Autonomous swarm coordination mode */
        attribute modeId :>> modeId = "MODE-004";
        attribute modeName :>> modeName = "Collaborative Operations";
        attribute description :>> description = "Multiple GARC vessels coordinate in swarm formation";
        attribute entryConditions :>> entryConditions = "Multiple vehicles initialized, swarm enabled";
        attribute exitConditions :>> exitConditions = "Mission complete, return to individual navigation";
    }
    
    requirement MODE_005 : SystemMode {
        doc /* Recovery and return mode */
        attribute modeId :>> modeId = "MODE-005";
        attribute modeName :>> modeName = "Recovery & Return";
        attribute description :>> description = "Vehicle returns to base/recovery point";
        attribute entryConditions :>> entryConditions = "Mission complete or abort commanded";
        attribute exitConditions :>> exitConditions = "At recovery point, safe to recover";
    }
}
"""
    
    FUNCTIONAL_FAILURES_PACKAGE = """
package FunctionalFailures {
    doc /* Functional failures for FHA */
    private import GARC_Model::*;
    private import SystemModes::*;
    
    // Functional failure metadata
    requirement def FunctionalFailure {
        attribute failureId : String;
        attribute system : String;
        attribute functionName : String;
        attribute failureMode : String;
        attribute failureDescription : String;
        attribute modePhase : String;
        attribute hazardCause ::> Hazards::Hazard;
        attribute linkedHazardId : String;
    }
    
    // Motion control failures
    requirement FF_MOT_001 : FunctionalFailure {
        doc /* Rudder control loss */
        attribute failureId :>> failureId = "FF-MOT-001";
        attribute system :>> system = "Motion Control";
        attribute functionName :>> functionName = "Rudder Actuation";
        attribute failureMode :>> failureMode = "Loss of Rudder Control";
        attribute failureDescription :>> failureDescription = "Rudder servo fails to respond to control commands";
        attribute modePhase :>> modePhase = "Navigation";
        attribute linkedHazardId :>> linkedHazardId = "HAZ-001";
    }
    
    requirement FF_COMM_001 : FunctionalFailure {
        doc /* C2 receiver failure */
        attribute failureId :>> failureId = "FF-COMM-001";
        attribute system :>> system = "Communication";
        attribute functionName :>> functionName = "Command Reception";
        attribute failureMode :>> failureMode = "C2 Receiver Failure";
        attribute failureDescription :>> failureDescription = "Radio receiver loses signal or malfunctions";
        attribute modePhase :>> modePhase = "AllPhases";
        attribute linkedHazardId :>> linkedHazardId = "HAZ-002";
    }
    
    requirement FF_PAYLOAD_001 : FunctionalFailure {
        doc /* Arming interlock failure */
        attribute failureId :>> failureId = "FF-PAYLOAD-001";
        attribute system :>> system = "Payload";
        attribute functionName :>> functionName = "Interlock Validation";
        attribute failureMode :>> failureMode = "Interlock Bypass";
        attribute failureDescription :>> failureDescription = "Safety interlocks fail to prevent arming in unsafe conditions";
        attribute modePhase :>> modePhase = "PayloadOperations";
        attribute linkedHazardId :>> linkedHazardId = "HAZ-PAYLOAD-001";
    }
}
"""
    
    @staticmethod
    def get_extension_sysml() -> str:
        """Get the SysML extension packages."""
        return SysMLModelExtender.SYSTEM_MODES_PACKAGE + "\n\n" + \
               SysMLModelExtender.FUNCTIONAL_FAILURES_PACKAGE


# ============================================================================
# FHA Excel Generator
# ============================================================================

class FHAExcelGenerator:
    """Generates FHA Excel worksheets from SysML model."""
    
    # Excel formatting constants
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    SUBHEADER_FILL = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    SUBHEADER_FONT = Font(bold=True, size=10)
    
    RISK_HIGH = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    RISK_MEDIUM = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    RISK_LOW = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    
    SEVERITY_MAP = {
        "catastrophic": "CAT I",
        "critical": "CAT II",
        "marginal": "CAT III",
        "negligible": "CAT IV",
    }
    
    LIKELIHOOD_MAP = {
        1e-7: "Remote",
        1e-5: "Occasional",
        1e-3: "Probable",
        1.0: "Frequent",
    }
    
    def __init__(self, template_path: Optional[str] = None):
        """Initialize FHA generator with optional template."""
        self.template_path = template_path
        self.wb = None
        self.ws = None
        self.current_row = 1
    
    def load_template(self) -> bool:
        """Load the FHA template or create new workbook."""
        if self.template_path and Path(self.template_path).exists():
            self.wb = openpyxl.load_workbook(self.template_path)
            self.ws = self.wb.active
            print(f"Loaded template: {self.template_path}")
            return True
        else:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.title = "FHA"
            print("Created new workbook")
            return False
    
    def create_fha_worksheet(self, analyzer: 'EnhancedFaultTreeAnalyzer') -> None:
        """Create comprehensive FHA worksheet."""
        
        # Setup worksheet
        self._setup_worksheet()
        
        # Add metadata
        self._add_metadata(analyzer)
        
        # Add headers
        self._add_headers()
        
        # Extract and add FHA rows
        fha_entries = self._extract_fha_entries(analyzer)
        self._add_fha_entries(fha_entries)
        
        # Format columns
        self._format_columns()
        
        # Add data validation and protection
        self._add_data_validation()
        
        print(f"Created FHA worksheet with {len(fha_entries)} entries")
    
    def _setup_worksheet(self) -> None:
        """Setup worksheet properties."""
        self.ws.page_setup.paperSize = self.ws.PAPERSIZE_LETTER
        self.ws.page_setup.orientation = 'landscape'
        self.ws.print_options.horizontalCentered = True
        self.current_row = 1
    
    def _add_metadata(self, analyzer: 'EnhancedFaultTreeAnalyzer') -> None:
        """Add document metadata."""
        
        # Title
        self.ws['A1'] = "FUNCTIONAL HAZARD ASSESSMENT (FHA)"
        self.ws['A1'].font = Font(bold=True, size=14)
        self.current_row = 2
        
        # Project info
        self.ws[f'A{self.current_row}'] = "Project:"
        self.ws[f'B{self.current_row}'] = "GARC Autonomous Vehicle"
        self.current_row += 1
        
        self.ws[f'A{self.current_row}'] = "Total Hazards:"
        self.ws[f'B{self.current_row}'] = len(analyzer.hazards)
        self.current_row += 1
        
        self.ws[f'A{self.current_row}'] = "Total Faults:"
        self.ws[f'B{self.current_row}'] = len(analyzer.faults)
        self.current_row += 1
        
        self.ws[f'A{self.current_row}'] = "Analysis Date:"
        from datetime import datetime
        self.ws[f'B{self.current_row}'] = datetime.now().strftime("%Y-%m-%d")
        self.current_row += 2
    
    def _add_headers(self) -> None:
        """Add column headers."""
        headers = [
            "ID",
            "Mode",
            "System",
            "Function Name",
            "Function Failure Mode",
            "Functional Failure Description",
            "Mode Phase",
            "Hazard/Effect",
            "Mishap",
            "Mitigations",
            "Severity",
            "Likelihood",
            "Risk",
            "SWCI",
            "SCC Condition",
            "SCC Rationale",
            "Severity Condition",
            "Reference Requirements",
            "Assumptions",
            "Notes/Questions"
        ]
        
        header_row = self.current_row
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=header_row, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Set header row height
        self.ws.row_dimensions[header_row].height = 30
        self.current_row += 1
    
    def _extract_fha_entries(self, analyzer: 'EnhancedFaultTreeAnalyzer') -> List[FunctionalFailure]:
        """Extract FHA entries from analyzer."""
        entries = []
        entry_id = 1
        
        # System modes
        system_modes = [
            SystemMode("MODE-001", "Startup", "System initialization"),
            SystemMode("MODE-002", "Navigation", "Vehicle transit to target"),
            SystemMode("MODE-003", "Payload Ops", "Payload preparation and arming"),
            SystemMode("MODE-004", "Collaboration", "Swarm coordination"),
            SystemMode("MODE-005", "Recovery", "Return to base"),
        ]
        
        # Create FHA entry for each fault
        for fault_id, fault in analyzer.faults.items():
            
            # Determine applicable modes
            applicable_modes = self._get_applicable_modes(fault, system_modes)
            
            for mode in applicable_modes:
                # Get hazard information
                hazard_id = fault.causes_hazards[0] if fault.causes_hazards else None
                hazard = analyzer.hazards.get(hazard_id) if hazard_id else None
                
                # Calculate risk
                likelihood = self._estimate_likelihood(fault)
                severity = hazard.severity if hazard else "negligible"
                risk = self._calculate_risk(severity, likelihood)
                
                entry = FunctionalFailure(
                    fha_id=f"FHA-{entry_id:03d}",
                    mode=mode.mode_name,
                    system=fault.component,
                    function_name=f"{fault.component} Operation",
                    failure_mode=fault.fault_id,
                    failure_description=fault.description,
                    mode_phase=mode.mode_id,
                    hazard_id=hazard_id or "",
                    hazard_effect=hazard.title if hazard else "",
                    mishap=hazard.effects if hazard else "",
                    mitigations=fault.mitigations,
                    severity=self.SEVERITY_MAP.get(severity.lower(), severity),
                    likelihood=likelihood,
                    risk=risk,
                    swci="Yes" if severity.lower() in ["catastrophic", "critical"] else "No",
                    scc_condition=f"Aircraft {hazard_id}" if hazard else "",
                    scc_rationale="Safety critical item due to high severity",
                    severity_condition=f"{severity.upper()} - {hazard.title}" if hazard else "",
                    reference_requirements=f"SR-{entry_id}",
                    assumptions="Standard operational environment",
                    notes=""
                )
                entries.append(entry)
                entry_id += 1
        
        return entries
    
    def _get_applicable_modes(self, fault: 'FaultNode', modes: List[SystemMode]) -> List[SystemMode]:
        """Determine which modes a fault is applicable to."""
        # For now, return all modes
        # In reality, this would be determined by the phase attribute
        return modes[:3]  # Default to first 3 modes
    
    def _estimate_likelihood(self, fault: 'FaultNode') -> str:
        """Estimate likelihood from failure rate."""
        if not fault.failure_rate:
            return "Unknown"
        
        if fault.failure_rate > 1e-3:
            return "Frequent"
        elif fault.failure_rate > 1e-5:
            return "Probable"
        elif fault.failure_rate > 1e-7:
            return "Occasional"
        else:
            return "Remote"
    
    def _calculate_risk(self, severity: str, likelihood: str) -> str:
        """Calculate risk level."""
        sev_map = {"catastrophic": 4, "critical": 3, "marginal": 2, "negligible": 1}
        like_map = {"Frequent": 4, "Probable": 3, "Occasional": 2, "Remote": 1}
        
        sev_score = sev_map.get(severity.lower(), 0)
        like_score = like_map.get(likelihood, 0)
        
        risk_score = sev_score * like_score
        
        if risk_score >= 12:
            return "Unacceptable"
        elif risk_score >= 6:
            return "High"
        elif risk_score >= 3:
            return "Medium"
        else:
            return "Low"
    
    def _add_fha_entries(self, entries: List[FunctionalFailure]) -> None:
        """Add FHA entries to worksheet."""
        for entry in entries:
            row = self.current_row
            
            # Convert entry to list
            values = [
                entry.fha_id,
                entry.mode,
                entry.system,
                entry.function_name,
                entry.failure_mode,
                entry.failure_description,
                entry.mode_phase,
                entry.hazard_effect,
                entry.mishap,
                entry.mitigations,
                entry.severity,
                entry.likelihood,
                entry.risk,
                entry.swci,
                entry.scc_condition,
                entry.scc_rationale,
                entry.severity_condition,
                entry.reference_requirements,
                entry.assumptions,
                entry.notes,
            ]
            
            # Write to cells
            for col, value in enumerate(values, 1):
                cell = self.ws.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Color code risk
                if col == 13:  # Risk column
                    if value == "Unacceptable":
                        cell.fill = self.RISK_HIGH
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif value == "High":
                        cell.fill = self.RISK_MEDIUM
                    elif value == "Low":
                        cell.fill = self.RISK_LOW
                
                # Bold SWCI items
                if col == 14 and value == "Yes":
                    cell.font = Font(bold=True)
            
            self.ws.row_dimensions[row].height = 40
            self.current_row += 1
    
    def _format_columns(self) -> None:
        """Format column widths and properties."""
        column_widths = {
            'A': 12,  # ID
            'B': 15,  # Mode
            'C': 15,  # System
            'D': 18,  # Function Name
            'E': 18,  # Failure Mode
            'F': 25,  # Failure Description
            'G': 15,  # Mode Phase
            'H': 20,  # Hazard/Effect
            'I': 20,  # Mishap
            'J': 20,  # Mitigations
            'K': 12,  # Severity
            'L': 12,  # Likelihood
            'M': 12,  # Risk
            'N': 8,   # SWCI
            'O': 15,  # SCC Condition
            'P': 20,  # SCC Rationale
            'Q': 18,  # Severity Condition
            'R': 20,  # Reference Req
            'S': 20,  # Assumptions
            'T': 20,  # Notes
        }
        
        for col, width in column_widths.items():
            self.ws.column_dimensions[col].width = width
        
        # Freeze header rows
        self.ws.freeze_panes = "A7"
    
    def _add_data_validation(self) -> None:
        """Add data validation and dropdown menus."""
        from openpyxl.worksheet.datavalidation import DataValidation
        
        # Severity dropdown
        dv_severity = DataValidation(
            type="list",
            formula1='"CAT I,CAT II,CAT III,CAT IV"',
            allow_blank=False
        )
        self.ws.add_data_validation(dv_severity)
        
        # Likelihood dropdown
        dv_likelihood = DataValidation(
            type="list",
            formula1='"Remote,Occasional,Probable,Frequent"',
            allow_blank=False
        )
        self.ws.add_data_validation(dv_likelihood)
        
        # Risk dropdown
        dv_risk = DataValidation(
            type="list",
            formula1='"Low,Medium,High,Unacceptable"',
            allow_blank=False
        )
        self.ws.add_data_validation(dv_risk)
        
        # SWCI dropdown
        dv_swci = DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        )
        self.ws.add_data_validation(dv_swci)
    
    def save(self, output_path: str) -> None:
        """Save the workbook."""
        self.wb.save(output_path)
        print(f"FHA worksheet saved to: {output_path}")


# ============================================================================
# Integration with Enhanced Analyzer
# ============================================================================

def generate_fha_excel(
    analyzer: 'EnhancedFaultTreeAnalyzer',
    template_path: Optional[str] = None,
    output_path: str = "GARC_FHA.xlsx"
) -> None:
    """Generate FHA Excel from analyzer."""
    
    print("\n" + "=" * 80)
    print("GENERATING FUNCTIONAL HAZARD ASSESSMENT (FHA)")
    print("=" * 80)
    
    generator = FHAExcelGenerator(template_path)
    generator.load_template()
    generator.create_fha_worksheet(analyzer)
    generator.save(output_path)
    
    print("FHA generation complete!")
